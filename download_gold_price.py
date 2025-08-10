import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook, Workbook
import os
import re

def unit_cn_to_en(unit_cn):
    """
    Convert Chinese unit to English unit.
    """
    table = {
        "元/克": "CNY/G",
        "元/千克": "CNY/KG",
        "元/吨": "CNY/ton",
        "元/盎司": "CNY/oz",
        "元/公斤": "CNY/KG",
        # 其他可能单位可以在这里补充
    }
    return table.get(unit_cn, unit_cn)  # 未知单位直接返回原文

def fetch_sge_gold_silver_prices_flat():
    """
    Fetches Shanghai Gold and Silver morning and afternoon benchmark prices from https://www.sge.com.cn/
    Returns a flat list: [date, metal, price_type, price, unit_en]
    Unit is parsed directly from the product description in HTML and translated to English.
    """
    url = "https://www.sge.com.cn/"
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/115.0.0.0 Safari/537.36"
    }
    response = requests.get(url, headers=headers)
    response.encoding = 'utf-8'
    if response.status_code != 200:
        print(f"Request failed, status code: {response.status_code}")
        return None

    soup = BeautifulSoup(response.text, "html.parser")

    # Helper: Parse price and unit from <li> block
    def parse_price_item(li):
        # Expect: <li><p>上海金午盘价（元/克）</p><span class="colorRed fs20">784.05</span></li>
        p_tag = li.find("p")
        price_span = li.find("span", class_="colorRed")
        if not p_tag or not price_span:
            return None, None
        # Extract unit from p_tag
        match = re.search(r"（(.*?)）", p_tag.get_text())
        unit_cn = match.group(1).strip() if match else ""
        unit_en = unit_cn_to_en(unit_cn)
        price = price_span.get_text(strip=True)
        return price, unit_en

    # Gold block
    gold_div = soup.find("div", id="dataStatistics0")
    if not gold_div:
        print("Cannot find Shanghai Gold data block.")
        return None
    gold_items = gold_div.find_all("li")
    if len(gold_items) < 3:
        print("Shanghai Gold data block structure error.")
        return None
    gold_date = gold_items[0].get_text(strip=True).replace("行情日期：", "")
    gold_am_price, gold_am_unit = parse_price_item(gold_items[1])
    gold_pm_price, gold_pm_unit = parse_price_item(gold_items[2])

    # Silver block
    silver_div = soup.find("div", id="dataStatistics1")
    if not silver_div:
        print("Cannot find Shanghai Silver data block.")
        return None
    silver_items = silver_div.find_all("li")
    if len(silver_items) < 3:
        print("Shanghai Silver data block structure error.")
        return None
    silver_date = silver_items[0].get_text(strip=True).replace("行情日期：", "")
    silver_am_price, silver_am_unit = parse_price_item(silver_items[1])
    silver_pm_price, silver_pm_unit = parse_price_item(silver_items[2])

    # Prepare flat data with EN unit
    data_rows = [
        [gold_date, "Gold", "AM", gold_am_price, gold_am_unit],
        [gold_date, "Gold", "PM", gold_pm_price, gold_pm_unit],
        [silver_date, "Silver", "AM", silver_am_price, silver_am_unit],
        [silver_date, "Silver", "PM", silver_pm_price, silver_pm_unit],
    ]
    return data_rows

def update_financial_data_gold_silver(excel_path="Financial_Data.xlsx", sheet_name="Precious_Metal_Prices"):
    """
    Appends new gold/silver price data to Financial_Data.xlsx in the 'Precious_Metal_Prices' sheet.
    Checks for duplicates before appending. If latest data already exists, prints info message.
    """
    new_rows = fetch_sge_gold_silver_prices_flat()
    if new_rows is None:
        print("No new price data found.")
        return

    # Check if file exists
    if os.path.exists(excel_path):
        wb = load_workbook(excel_path)
        # If sheet exists, use it, else create
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(sheet_name)
            ws.append(["date", "metal", "price_type", "price", "unit"])  # Add unit column
        # Read all existing rows (skip header)
        existing_rows = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            existing_rows.add(tuple(row))
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        ws.append(["date", "metal", "price_type", "price", "unit"])  # Add unit column
        existing_rows = set()

    # Check if all new rows are already in the sheet
    is_latest = True
    for new_row in new_rows:
        if tuple(new_row) not in existing_rows:
            is_latest = False
            ws.append(new_row)

    if is_latest:
        print("Excel file is already up-to-date with the latest precious metal prices.")
    else:
        wb.save(excel_path)
        print(f"New precious metal price data appended to Excel file: {excel_path}, sheet: {sheet_name}")

# 示例调用
if __name__ == "__main__":
    update_financial_data_gold_silver()