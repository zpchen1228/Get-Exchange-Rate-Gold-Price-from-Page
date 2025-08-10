from openpyxl import load_workbook

def get_latest_rates(file_path, currencies, sheet_name="Exchange_Rate"):
    """
    Get the latest rates for given currencies from the bottom of the specified Excel sheet.
    :param file_path: Excel file path
    :param currencies: List of currency codes, e.g. ["USD", "EUR"]
    :param sheet_name: Sheet name to read from (default: "Exchange_Rate")
    :return: Dictionary {currency_code: rate}
    """
    rates = {}
    try:
        wb = load_workbook(file_path)
        ws = wb[sheet_name]
        last_row = ws.max_row

        # Traverse from last row upwards (skip header)
        for row in range(last_row, 1, -1):
            currency_name = ws.cell(row, 2).value
            if currency_name is None:
                continue
            currency_name = currency_name.strip()
            for currency_code in currencies:
                if currency_code in currency_name:
                    rates[currency_code] = ws.cell(row, 3).value
                    if len(rates) == len(currencies):
                        return rates
    except Exception as e:
        print(f"Error reading rates: {e}")

    return rates