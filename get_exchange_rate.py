import os
import requests
from openpyxl import Workbook, load_workbook
from datetime import datetime

def download_exchange_rate_data():
    """Download exchange rate data from PBOC and save to 'Financial_Data.xlsx' in 'Exchange_Rate' sheet."""
    # File and sheet settings
    save_dir = os.path.dirname(os.path.abspath(__file__))
    file_path = os.path.join(save_dir, "Financial_Data.xlsx")
    sheet_name = "Exchange_Rate"

    # Ensure directory exists
    os.makedirs(save_dir, exist_ok=True)

    # 1. Check if file and sheet exist, get latest date from sheet if possible
    existing_date = None
    if os.path.exists(file_path):
        wb = load_workbook(file_path)
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            last_row = ws.max_row
            if last_row > 1:
                existing_date = ws.cell(last_row, 4).value
        else:
            ws = wb.create_sheet(sheet_name)
            # Write headers
            ws.cell(1, 1).value = "Id"
            ws.cell(1, 2).value = "Currency"
            ws.cell(1, 3).value = "Exchange Rate"
            ws.cell(1, 4).value = "Date"
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        # Write headers
        ws.cell(1, 1).value = "Id"
        ws.cell(1, 2).value = "Currency"
        ws.cell(1, 3).value = "Exchange Rate"
        ws.cell(1, 4).value = "Date"

    # 2. Get latest date from API
    try:
        url = "https://www.chinamoney.com.cn/r/cms/www/chinamoney/data/fx/ccpr.json"
        res = requests.get(url, headers={"User-Agent": "Mozilla/5.0"}, timeout=10)
        res.raise_for_status()
        api_date = res.json()["data"]["lastDate"]

        # 3. Check if data is already up-to-date
        if existing_date == api_date:
            print("This data is the latest.")
            return
    except Exception as e:
        print(f"Error getting data: {e}")
        return

    # 4. Get full data from API
    data = res.json()
    data_list = data["records"]

    # 5. Prepare Excel sheet for appending
    start_row = ws.max_row + 1 if ws.max_row > 1 else 2

    # 6. Append new data
    for idx, item in enumerate(data_list, start=start_row):
        unique_id = f"{api_date}-{item['vrtCode']}-{datetime.now().strftime('%H%M%S%f')}"
        ws.cell(idx, 1).value = unique_id
        ws.cell(idx, 2).value = item["vrtEName"]
        ws.cell(idx, 3).value = item["price"]
        ws.cell(idx, 4).value = api_date

    # 7. Save file
    wb.save(file_path)
    print(f"Successfully saved {len(data_list)} records to: {file_path}, sheet: {sheet_name}")

if __name__ == "__main__":
    download_exchange_rate_data()