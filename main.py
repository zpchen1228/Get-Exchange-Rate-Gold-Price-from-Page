from get_exchange_rate import download_exchange_rate_data
from send_email import send_email
from get_data_from_excel import get_latest_rates
from download_gold_price import fetch_sge_gold_silver_prices_flat
import schedule
import time
import datetime
import pytz
import os

def main_task():
    """Main task function: download exchange rate and gold/silver prices, send email report (PM price only, optimized title/date placement)"""
    print(f"{datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')} - Starting exchange rate and gold/silver price email task")

    # Use Financial_Data.xlsx and Exchange_Rate sheet
    save_dir = os.path.dirname(os.path.abspath(__file__))
    financial_file = os.path.join(save_dir, "Financial_Data.xlsx")
    download_exchange_rate_data()  # Assumes already writes to correct file & sheet

    # Get latest USD and EUR rates from the Excel file
    currencies = ["USD", "EUR"]
    latest_rates = get_latest_rates(financial_file, currencies, sheet_name="Exchange_Rate")

    # 获取最新一行的汇率日期信息（Exchange_Rate sheet, 第4列，底部最后一行）
    rate_date = None
    try:
        from openpyxl import load_workbook
        wb = load_workbook(financial_file, read_only=True)
        ws = wb["Exchange_Rate"]
        last_row = ws.max_row
        rate_date = ws.cell(last_row, 4).value
    except Exception as e:
        print(f"Error reading exchange rate date: {e}")
        rate_date = None

    # 汇率信息格式化（不含日期）
    rates_html = ""
    for currency_code in currencies:
        if currency_code in latest_rates:
            rate = latest_rates[currency_code]
            currency_name = "USD" if currency_code == "USD" else "EUR"
            rates_html += f"<li>{currency_name}/CNY: <b>{rate}</b></li>"
        else:
            rates_html += f"<li>{currency_code} rate not found</li>"

    # 获取贵金属价格，只展示午盘价，同时获取贵金属价格日期
    # 修改贵金属价格处理逻辑（原代码第45-60行）
    gold_silver_rows = fetch_sge_gold_silver_prices_flat()  # 返回格式应为 [[date, metal, price_type, price, unit], ...]
    gold_silver_pm_html = ""
    metal_date = None
    if gold_silver_rows:
        pm_rows = [row for row in gold_silver_rows if row[2] == "PM"]
        if pm_rows:
            metal_date = pm_rows[0][0]
            for row in pm_rows:
                # 更新解包逻辑，包含unit列
                _, metal, _, price, unit = row  # 解包5个值[3,6](@ref)
                metal_en = "Gold" if metal == "Gold" else "Silver"
                gold_silver_pm_html += f"<li>{metal_en}/CNY PM Price: <b>{price} {unit}</b></li>"  # 添加单位显示[1](@ref)
        else:
            gold_silver_pm_html = "<li>No gold or silver PM price data found</li>"
    else:
        gold_silver_pm_html = "<li>No gold or silver PM price data found</li>"

    # 邮件正文
    email_content = f"""
    <html>
        <body>
            <p><b>Dear Zhi Ping,</b></p>
            <p>Please find today's foreign exchange rate and precious metal PM price report attached.</p>
            <p><b>Latest Key Currency Rates{' (' + str(rate_date) + ')' if rate_date else ''}:</b></p>
            <ul>
                {rates_html}
            </ul>
            <p><b>Latest Precious Metal PM Prices{' (' + str(metal_date) + ')' if metal_date else ''}:</b></p>
            <ul>
                {gold_silver_pm_html}
            </ul>
            <p><b>Report Highlights:</b></p>
            <ul>
                <li>Rates are the most recent entries from the People's Bank of China</li>
                <li>Gold and silver PM prices are official Shanghai Gold Exchange benchmark prices</li>
                <li>Full data set available in the attached Excel file</li>
            </ul>
            <p>Please let me know if you need additional analysis or specific currency pairs.</p>
            <br>
            <p><b>Best Regards,</b></p>
            <p>Chen Zhi Ping<br>
            Financial Data Analyst</p>
        </body>
    </html>
    """

    # Send email
    receive_address = [""]
    receive_address_str = ", ".join(receive_address)

    try:
        send_email(
            sender_email="",
            sender_name="Chen Zhi Ping",
            sender_auth="",
            receiver_email=receive_address_str,
            subject="Daily Exchange Rate & Gold/Silver PM Price Report",
            content=email_content,
            content_type="html",
            attachment_path=financial_file
        )
        print(f"{datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')} - Email sent successfully")
    except Exception as e:
        print(f"{datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')} - Failed to send email: {str(e)}")

if __name__ == "__main__":
    # Set China timezone
    china_tz = pytz.timezone('Asia/Shanghai')
    # Run once at startup
    print(f"{datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')} - Program started, first task running")
    main_task()
    # Schedule daily at 09:30 China time
    schedule.every().day.at("09:30", china_tz).do(main_task)
    print(f"Scheduler started. Next run: {schedule.next_run().astimezone(china_tz).strftime('%Y-%m-%d %H:%M:%S')}")
    # Main loop
    while True:
        schedule.run_pending()
        time.sleep(60)